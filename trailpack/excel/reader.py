"""Excel reader module for loading and inspecting Excel files.

This module provides an ExcelReader class that:
- Loads only the structure (sheets and columns) into memory
- Provides access to sheet names
- Provides access to column names for mapping
"""

from pathlib import Path
from typing import Dict, List, Optional, Union
import openpyxl


class ExcelReader:
    """
    Excel file reader that loads only sheet structure (sheets and columns) into memory.

    This class provides methods to inspect Excel file structure without loading
    all the data, making it memory-efficient for large files:
    - List available sheets
    - Get column names from specific sheets

    Example:
        >>> reader = ExcelReader("data.xlsx")
        >>> sheet_names = reader.sheets()
        >>> columns = reader.columns("Sheet1")
    """

    def __init__(self, file_path: Union[str, Path], header_row: int = 1):
        """
        Initialize ExcelReader and load sheet structure (sheets and columns) into memory.

        Only the sheet names and column headers are loaded, not the actual data.
        This makes it memory-efficient for large Excel files.

        Args:
            file_path: Path to the Excel file (.xlsx, .xlsm, .xltx, .xltm)
            header_row: Row number containing column headers (1-indexed). Defaults to 1.

        Raises:
            FileNotFoundError: If the file does not exist
            ValueError: If the file is not a valid Excel file
        """
        self.file_path = Path(file_path)
        self.header_row = header_row

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        if self.file_path.suffix not in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            raise ValueError(f"File must be an Excel file (.xlsx, .xlsm, .xltx, .xltm), got: {self.file_path.suffix}")

        # Load only sheet structure (sheets -> columns) into memory
        self._sheet_columns: Dict[str, List[str]] = {}
        self._load_structure()

    def _load_structure(self):
        """
        Load sheet structure (sheet names and column headers) from Excel file.

        Opens the workbook in read-only mode, extracts structure, then closes it.
        Only loads metadata, not actual data, for memory efficiency.
        """
        try:
            # Open in read-only mode for efficiency
            workbook = openpyxl.load_workbook(
                self.file_path,
                read_only=True,
                data_only=True
            )

            # Extract sheet names and columns
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Read header row to get column names
                columns = []
                try:
                    # Get the header row
                    for row in sheet.iter_rows(min_row=self.header_row, max_row=self.header_row):
                        for cell in row:
                            # Convert cell value to string, use empty string for None
                            column_name = str(cell.value) if cell.value is not None else ""
                            columns.append(column_name)
                        break  # Only read the header row

                    # Remove trailing empty columns
                    while columns and columns[-1] == "":
                        columns.pop()

                except Exception as e:
                    # If we can't read columns (e.g., empty sheet), store empty list
                    columns = []

                self._sheet_columns[sheet_name] = columns

            # Close workbook immediately - we only need the structure
            workbook.close()

        except Exception as e:
            raise ValueError(f"Failed to load Excel file structure: {e}")

    def sheets(self) -> List[str]:
        """
        Get list of all sheet names in the workbook.

        Returns:
            List of sheet names as strings

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> sheet_names = reader.sheets()
            >>> print(sheet_names)
            ['Sheet1', 'Sheet2', 'Data']
        """
        return list(self._sheet_columns.keys())

    def columns(self, sheet_name: Optional[str] = None) -> List[str]:
        """
        Get list of column names from a specific sheet.

        Args:
            sheet_name: Name of the sheet to read columns from.
                       If None, uses the first sheet.

        Returns:
            List of column names as strings. Empty cells are returned as empty strings.

        Raises:
            ValueError: If sheet_name doesn't exist in the workbook

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> columns = reader.columns("Sheet1")
            >>> print(columns)
            ['ID', 'Name', 'Value', 'Date']
        """
        # Get sheet name
        if sheet_name is None:
            # Use first sheet if no name specified
            sheet_name = self.sheets()[0] if self.sheets() else None
            if sheet_name is None:
                return []

        # Check if sheet exists
        if sheet_name not in self._sheet_columns:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {', '.join(self.sheets())}"
            )

        return self._sheet_columns[sheet_name]

    def get_structure(self) -> Dict[str, List[str]]:
        """
        Get the complete sheet structure as a dictionary.

        Returns:
            Dictionary mapping sheet names to their column lists

        Example:
            >>> reader = ExcelReader("data.xlsx")
            >>> structure = reader.get_structure()
            >>> print(structure)
            {'Sheet1': ['ID', 'Name', 'Value'], 'Sheet2': ['Date', 'Amount']}
        """
        return self._sheet_columns.copy()

    def reload(self):
        """
        Reload the sheet structure from the Excel file.

        Useful if the file has been modified and you want to refresh the structure.
        """
        self._sheet_columns.clear()
        self._load_structure()

    def __enter__(self):
        """Context manager entry."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - nothing to clean up as workbook is already closed."""
        pass

    def __repr__(self) -> str:
        """String representation of ExcelReader."""
        return f"ExcelReader(file_path='{self.file_path}', sheets={len(self.sheets())})"
